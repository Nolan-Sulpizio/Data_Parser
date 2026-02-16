#!/usr/bin/env python3
"""
Wesco MRO Data Parser — Comprehensive Validation Suite
======================================================
Purpose: Validate parser accuracy, performance, and QA engine against
         production datasets for team confidence and Gamma presentation.

Usage:
    python tests/run_validation.py
    python tests/run_validation.py --data-dir "/custom/path"

Author: Nolan Sulpizio / Wesco Global Accounts
"""

import os
import sys
import re
import time
import json
import argparse
from datetime import datetime
from pathlib import Path
import pandas as pd
import traceback

# Add parent directory to path to import engine
sys.path.insert(0, str(Path(__file__).parent.parent))
from engine.parser_core import (
    pipeline_mfg_pn, extract_pn_from_text, extract_mfg_from_text,
    sanitize_mfg, NORMALIZE_MFG, DISTRIBUTORS, DESCRIPTORS
)

# ═══════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════

DEFAULT_DATA_DIR = "/Users/nolansulpizio/Desktop/Documents - Nolan's MacBook Pro/WESCO/Data Parse Agent/Data Context"

# Test file names
ELECTRICAL_BLANK    = "Electrical PN_MFG Search.XLSX"
ELECTRICAL_COMPLETE = "Electrical PN_MFG Search_COMPLETE.xlsx"
BOOK25              = "Book25.xlsx"

# ═══════════════════════════════════════════════════════════════
# TEST LOGGER
# ═══════════════════════════════════════════════════════════════

class TestLogger:
    """Structured test logging with console and file output."""

    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.log_path = self.output_dir / "test_log.txt"
        self.entries = []
        self._write_header()

    def _write_header(self):
        self._write("=" * 70)
        self._write("WESCO MRO DATA PARSER — VALIDATION TEST RUN")
        self._write(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._write("=" * 70)
        self._write("")

    def _write(self, msg):
        print(msg)
        with open(self.log_path, 'a') as f:
            f.write(msg + '\n')

    def section(self, title):
        self._write("")
        self._write("─" * 60)
        self._write(f"  {title}")
        self._write("─" * 60)

    def result(self, test_name, passed, detail=""):
        status = "✓ PASS" if passed else "✗ FAIL"
        msg = f"  [{status}]  {test_name}"
        if detail:
            msg += f"  →  {detail}"
        self._write(msg)
        self.entries.append({
            'test': test_name,
            'passed': passed,
            'detail': detail,
            'timestamp': datetime.now().isoformat()
        })

    def info(self, msg):
        self._write(f"  ℹ  {msg}")

    def warning(self, msg):
        self._write(f"  ⚠  {msg}")

    def summary(self):
        total = len(self.entries)
        passed = sum(1 for e in self.entries if e['passed'])
        failed = total - passed
        pass_rate = round(passed / total * 100, 1) if total > 0 else 0

        self._write("")
        self._write("=" * 70)
        self._write(f"  TEST SUMMARY: {passed}/{total} passed ({pass_rate}%)")
        if failed > 0:
            self._write(f"  ⚠ {failed} tests failed - review output above")
        else:
            self._write("  ✓ All tests passed!")
        self._write("=" * 70)
        return passed, failed

# ═══════════════════════════════════════════════════════════════
# TEST PHASE 1: ACCURACY — Ground Truth Comparison
# ═══════════════════════════════════════════════════════════════

def test_accuracy(data_dir, output_dir, log):
    """Compare parser output against manually-verified ground truth."""
    log.section("PHASE 1: ACCURACY — Electrical PN/MFG Extraction vs Ground Truth")

    blank_path = Path(data_dir) / ELECTRICAL_BLANK
    truth_path = Path(data_dir) / ELECTRICAL_COMPLETE

    # Verify files exist
    if not blank_path.exists():
        log.result("Blank file availability", False, f"Missing: {blank_path}")
        return None
    if not truth_path.exists():
        log.result("Truth file availability", False, f"Missing: {truth_path}")
        return None

    log.info(f"Input:  {blank_path}")
    log.info(f"Truth:  {truth_path}")

    # Parse the blank file
    parsed_output = Path(output_dir) / "electrical_parsed_output.xlsx"
    start = time.time()
    try:
        df_blank = pd.read_excel(blank_path, engine='openpyxl')
        df_blank.columns = [c.strip() for c in df_blank.columns]

        # Run parser
        source_cols = [c for c in df_blank.columns if 'description' in c.lower() or 'text' in c.lower()]
        result = pipeline_mfg_pn(df_blank, source_cols=source_cols, mfg_col='MFG', pn_col='PN', add_sim=True)
        parsed_df = result.df

        # Save output
        parsed_df.to_excel(parsed_output, index=False, engine='openpyxl')

    except Exception as e:
        log.result("Parser execution", False, f"CRASHED: {e}")
        traceback.print_exc()
        return None

    elapsed = time.time() - start
    log.result("Parser execution", True, f"Completed in {elapsed:.2f}s")

    # Load ground truth
    truth_df = pd.read_excel(truth_path, engine='openpyxl')
    truth_df.columns = [c.strip() for c in truth_df.columns]

    # Compare results
    results = {
        'total_rows': len(truth_df),
        'elapsed_seconds': round(elapsed, 2),
        'mfg': {'exact_match': 0, 'fuzzy_match': 0, 'mismatch': 0, 'both_blank': 0, 'truth_blank': 0},
        'pn':  {'exact_match': 0, 'fuzzy_match': 0, 'mismatch': 0, 'both_blank': 0, 'truth_blank': 0},
        'mismatches': []
    }

    for idx in range(min(len(truth_df), len(parsed_df))):
        for field in ['MFG', 'PN']:
            truth_val = str(truth_df.at[idx, field]).strip().upper() if pd.notna(truth_df.at[idx, field]) else ''
            parsed_val = str(parsed_df.at[idx, field]).strip().upper() if pd.notna(parsed_df.at[idx, field]) else ''

            # Normalize spaces for comparison
            truth_clean = re.sub(r'\s+', '', truth_val)
            parsed_clean = re.sub(r'\s+', '', parsed_val)

            if truth_clean == '' and parsed_clean == '':
                results[field.lower()]['both_blank'] += 1
            elif truth_clean == '':
                results[field.lower()]['truth_blank'] += 1
            elif truth_clean == parsed_clean:
                results[field.lower()]['exact_match'] += 1
            elif truth_clean in parsed_clean or parsed_clean in truth_clean:
                results[field.lower()]['fuzzy_match'] += 1
            else:
                results[field.lower()]['mismatch'] += 1
                desc = str(truth_df.at[idx, 'Material Description'])[:60] if 'Material Description' in truth_df.columns else ''
                results['mismatches'].append({
                    'row': idx + 2,  # Excel row (1-indexed + header)
                    'field': field,
                    'expected': truth_val,
                    'got': parsed_val,
                    'description': desc
                })

    # Calculate accuracy metrics
    for field in ['mfg', 'pn']:
        r = results[field]
        scorable = r['exact_match'] + r['fuzzy_match'] + r['mismatch']
        r['scorable_rows'] = scorable
        r['accuracy_exact'] = round(r['exact_match'] / scorable * 100, 1) if scorable > 0 else 0
        r['accuracy_fuzzy'] = round((r['exact_match'] + r['fuzzy_match']) / scorable * 100, 1) if scorable > 0 else 0

    # Log results
    for field, label in [('mfg', 'MFG'), ('pn', 'PN')]:
        r = results[field]
        log.result(
            f"{label} Exact Match Rate",
            r['accuracy_exact'] >= 85,
            f"{r['accuracy_exact']}% ({r['exact_match']}/{r['scorable_rows']} rows)"
        )
        log.result(
            f"{label} Fuzzy Match Rate",
            r['accuracy_fuzzy'] >= 92,
            f"{r['accuracy_fuzzy']}% ({r['exact_match'] + r['fuzzy_match']}/{r['scorable_rows']} rows)"
        )
        log.info(f"{label} Stats: {r['mismatch']} mismatches | {r['both_blank']} both blank | {r['truth_blank']} truth blank")

    log.result("Processing Time", elapsed < 2.0, f"{elapsed:.2f}s (target: <2s)")
    log.result("Total Mismatches", len(results['mismatches']) < 25, f"{len(results['mismatches'])} found (target: <25)")

    return results

# ═══════════════════════════════════════════════════════════════
# TEST PHASE 2: QA ENGINE — Flag Detection
# ═══════════════════════════════════════════════════════════════

def test_qa_engine(output_dir, log):
    """Validate QA engine catches data quality issues."""
    log.section("PHASE 2: QA ENGINE — Validation Rule Detection")

    parsed_path = Path(output_dir) / "electrical_parsed_output.xlsx"
    if not parsed_path.exists():
        log.result("QA input file", False, "Parsed output not found (run accuracy test first)")
        return None

    df = pd.read_excel(parsed_path, engine='openpyxl')
    df.columns = [c.strip() for c in df.columns]

    # Define QA rules
    QA_RULES = [
        ('MFG_missing',        lambda r: str(r.get('MFG', '')).strip() in ['', 'nan', 'None']),
        ('PN_missing',         lambda r: str(r.get('PN', '')).strip() in ['', 'nan', 'None']),
        ('MFG_is_distributor', lambda r: str(r.get('MFG', '')).upper().strip() in DISTRIBUTORS),
        ('MFG_is_descriptor',  lambda r: str(r.get('MFG', '')).upper().strip() in DESCRIPTORS),
        ('MFG_has_digits',     lambda r: bool(re.search(r"\d", str(r.get('MFG', '')))) and str(r.get('MFG', '')).strip() not in ['', 'nan']),
        ('CROUSE_hyphenated',  lambda r: 'CROUSE-HINDS' in str(r.get('MFG', '')).upper()),
        ('SQUARE_D_variant',   lambda r: str(r.get('MFG', '')).upper().strip() == 'SQUARE-D'),
    ]

    issues = []
    flag_counts = {}

    for idx, row in df.iterrows():
        for key, fn in QA_RULES:
            try:
                if fn(row):
                    issues.append({
                        'row': idx + 2,
                        'flag': key,
                        'MFG': row.get('MFG', ''),
                        'PN': row.get('PN', ''),
                        'description': str(row.get('Material Description', ''))[:60]
                    })
                    flag_counts[key] = flag_counts.get(key, 0) + 1
            except Exception:
                pass

    results = {
        'total_rows': len(df),
        'total_flags': len(issues),
        'flag_counts': flag_counts,
        'issues': issues
    }

    # Log results
    log.result("QA engine execution", True, f"Scanned {len(df)} rows")
    log.result("No distributors in MFG", flag_counts.get('MFG_is_distributor', 0) == 0, f"{flag_counts.get('MFG_is_distributor', 0)} found")
    log.result("No descriptors in MFG", flag_counts.get('MFG_is_descriptor', 0) == 0, f"{flag_counts.get('MFG_is_descriptor', 0)} found")
    log.result("No digits in MFG", flag_counts.get('MFG_has_digits', 0) == 0, f"{flag_counts.get('MFG_has_digits', 0)} found")
    log.result("CROUSE HINDS normalized", flag_counts.get('CROUSE_hyphenated', 0) == 0, f"{flag_counts.get('CROUSE_hyphenated', 0)} hyphenated")
    log.result("SQUARE D normalized", flag_counts.get('SQUARE_D_variant', 0) == 0, f"{flag_counts.get('SQUARE_D_variant', 0)} variants")

    log.info(f"Total QA flags raised: {len(issues)}")
    for flag, count in sorted(flag_counts.items(), key=lambda x: -x[1])[:5]:
        log.info(f"  {flag}: {count}")

    return results

# ═══════════════════════════════════════════════════════════════
# TEST PHASE 3: UNIT TESTS — Individual Functions
# ═══════════════════════════════════════════════════════════════

def test_unit_extraction(log):
    """Test core extraction functions with known inputs/outputs."""
    log.section("PHASE 3: UNIT TESTS — Core Function Validation")

    # PN extraction test cases
    pn_cases = [
        ("MARKER,WIRE MANUFACTURER: PANDUIT PART NUMBER: PCMB-8", "PCMB-8", "labeled PART NUMBER:"),
        ("SENSOR,PROX,USG CORP,PN: 10690735", "10690735", "labeled PN:"),
        ("FUSE,NM,SHAWMUT,MN: AJT250,250A,TYPE J", "AJT250", "labeled MN:"),
        ("MODEL NUMBER: 6030228 PN: 10690735", "10690735", "PN: priority over MODEL NUMBER:"),
        ("VALVE,BALL,3/4IN,SHARPE,VA123", "VA123", "heuristic fallback (letters+digits)"),
    ]

    pn_pass = 0
    for text, expected, desc in pn_cases:
        result, method = extract_pn_from_text(text)
        match = (result or '').strip().upper() == (expected or '').strip().upper()
        log.result(f"PN: {desc}", match, f"expected={expected}, got={result} [{method}]")
        if match:
            pn_pass += 1

    # MFG extraction test cases
    mfg_cases = [
        ("MANUFACTURER: PANDUIT PART NUMBER: PCMB-8", "PANDUIT", "labeled MANUFACTURER:"),
        ("MANUFACTURER: FERRAZ SHAWMUT MODEL NUMBER: AJT250", "FERRAZ SHAWMUT", "multi-word MFG"),
        ("MANUFACTURER: USG CORPORATION MODEL NUMBER: 6030228", "USG CORPORATION", "MFG with CORPORATION"),
    ]

    mfg_pass = 0
    for text, expected, desc in mfg_cases:
        result, method = extract_mfg_from_text(text, None, set())
        result_norm = re.sub(r'\s+', ' ', result.strip().upper()) if result else None
        expected_norm = expected.strip().upper() if expected else None
        match = result_norm == expected_norm
        log.result(f"MFG: {desc}", match, f"expected={expected}, got={result} [{method}]")
        if match:
            mfg_pass += 1

    # Sanitization test cases
    sanitize_cases = [
        ("PANDUIT", "PANDUIT", "valid MFG passes through"),
        ("GRAYBAR", None, "distributor rejected"),
        ("SENSOR123", None, "digits rejected"),
        ("PANDT", "PANDUIT", "abbreviation normalized"),
        ("CROUSE-HINDS", "CROUSE HINDS", "hyphen removed"),
        ("STATIC O RING", "STATIC O-RING", "canonical form"),
        ("SWITCH ASSEMBLY", None, "descriptor keyword rejected"),
    ]

    san_pass = 0
    for input_val, expected, desc in sanitize_cases:
        result = sanitize_mfg(input_val)
        match = result == expected
        log.result(f"Sanitize: {desc}", match, f"expected={expected}, got={result}")
        if match:
            san_pass += 1

    return {
        'pn_tests': len(pn_cases),
        'pn_passed': pn_pass,
        'mfg_tests': len(mfg_cases),
        'mfg_passed': mfg_pass,
        'sanitize_tests': len(sanitize_cases),
        'sanitize_passed': san_pass
    }

# ═══════════════════════════════════════════════════════════════
# TEST PHASE 4: SCALE TEST — Book25 (174 rows)
# ═══════════════════════════════════════════════════════════════

def test_scale_book25(data_dir, output_dir, log):
    """Test parser performance and extraction on large, differently-formatted dataset."""
    log.section("PHASE 4: SCALE TEST — Book25 (174 rows, different schema)")

    book25_path = Path(data_dir) / BOOK25
    if not book25_path.exists():
        log.result("Book25 file availability", False, f"Missing: {book25_path}")
        return None

    # Load and adapt Book25
    try:
        df_raw = pd.read_excel(book25_path, engine='openpyxl')
        df_raw.columns = [c.strip() for c in df_raw.columns]
        log.info(f"Book25 loaded: {df_raw.shape}")

        # Adapt to parser format
        desc_cols = [c for c in df_raw.columns if 'MATERIAL DESCRIPTION' in c.upper() or 'DESCRIPTION' in c.upper()]
        adapted = pd.DataFrame()
        if desc_cols:
            adapted['Material Description'] = df_raw[desc_cols[0]]
            if len(desc_cols) > 1:
                adapted['Material PO Text'] = df_raw[desc_cols[1]]
        adapted['MFG'] = pd.Series([''] * len(df_raw), dtype='object')
        adapted['PN'] = pd.Series([''] * len(df_raw), dtype='object')

        adapted_path = Path(output_dir) / "book25_adapted_input.xlsx"
        adapted.to_excel(adapted_path, index=False, engine='openpyxl')

    except Exception as e:
        log.result("Book25 load and adapt", False, f"CRASHED: {e}")
        traceback.print_exc()
        return None

    # Parse Book25
    parsed_path = Path(output_dir) / "book25_parsed_output.xlsx"
    start = time.time()
    try:
        source_cols = [c for c in adapted.columns if 'description' in c.lower() or 'text' in c.lower()]
        result = pipeline_mfg_pn(adapted, source_cols=source_cols, mfg_col='MFG', pn_col='PN', add_sim=True)
        parsed_df = result.df
        parsed_df.to_excel(parsed_path, index=False, engine='openpyxl')

    except Exception as e:
        log.result("Book25 parser execution", False, f"CRASHED: {e}")
        traceback.print_exc()
        return None

    elapsed = time.time() - start
    log.result("Book25 parser execution", True, f"174 rows in {elapsed:.2f}s")
    log.result("Performance threshold", elapsed < 5.0, f"{elapsed:.2f}s (target: <5s)")

    # Compare against ground truth if available
    if 'Manufacturer 1' in df_raw.columns and 'Part Number 1' in df_raw.columns:
        truth_mfg = df_raw['Manufacturer 1'].fillna('').astype(str).str.strip().str.upper()
        truth_pn = df_raw['Part Number 1'].fillna('').astype(str).str.strip().str.upper()
        parsed_mfg = parsed_df['MFG'].fillna('').astype(str).str.strip().str.upper()
        parsed_pn = parsed_df['PN'].fillna('').astype(str).str.strip().str.upper()

        # MFG matching
        mfg_exact = mfg_fuzzy = mfg_total = 0
        for i in range(len(truth_mfg)):
            t = truth_mfg.iloc[i].strip()
            p = parsed_mfg.iloc[i].strip()
            if t == '':
                continue
            mfg_total += 1
            t_clean = re.sub(r'[\s\-]', '', t)
            p_clean = re.sub(r'[\s\-]', '', p)
            if t_clean == p_clean:
                mfg_exact += 1
            elif p and (t_clean in p_clean or p_clean in t_clean):
                mfg_fuzzy += 1

        # PN matching
        pn_exact = pn_fuzzy = pn_total = 0
        for i in range(len(truth_pn)):
            t = re.sub(r'\s+', '', truth_pn.iloc[i].strip())
            p = re.sub(r'\s+', '', parsed_pn.iloc[i].strip())
            if t == '':
                continue
            pn_total += 1
            if t == p:
                pn_exact += 1
            elif t in p or p in t:
                pn_fuzzy += 1

        mfg_rate = round(mfg_exact / mfg_total * 100, 1) if mfg_total > 0 else 0
        pn_rate = round(pn_exact / pn_total * 100, 1) if pn_total > 0 else 0

        log.result("Book25 MFG extraction", mfg_rate > 25, f"{mfg_rate}% exact ({mfg_exact}/{mfg_total}), {mfg_fuzzy} fuzzy")
        log.result("Book25 PN extraction", pn_rate > 60, f"{pn_rate}% exact ({pn_exact}/{pn_total}), {pn_fuzzy} fuzzy")

        mfg_filled = (parsed_mfg.str.strip() != '').sum()
        pn_filled = (parsed_pn.str.strip() != '').sum()
        log.info(f"MFG populated: {mfg_filled}/{len(parsed_df)} rows ({round(mfg_filled/len(parsed_df)*100,1)}%)")
        log.info(f"PN populated: {pn_filled}/{len(parsed_df)} rows ({round(pn_filled/len(parsed_df)*100,1)}%)")

        return {
            'total_rows': len(parsed_df),
            'elapsed_seconds': round(elapsed, 2),
            'mfg_exact': mfg_exact,
            'mfg_fuzzy': mfg_fuzzy,
            'mfg_total': mfg_total,
            'pn_exact': pn_exact,
            'pn_fuzzy': pn_fuzzy,
            'pn_total': pn_total
        }
    else:
        log.warning("Book25 ground truth columns not found - skipping accuracy comparison")
        return {'total_rows': len(parsed_df), 'elapsed_seconds': round(elapsed, 2)}

# ═══════════════════════════════════════════════════════════════
# TEST PHASE 5: NORMALIZATION — Data Quality
# ═══════════════════════════════════════════════════════════════

def test_normalization(log):
    """Validate normalization rules and data quality mappings."""
    log.section("PHASE 5: NORMALIZATION — MFG Name Standardization")

    # Check for circular references
    circular = []
    for abbrev, canonical in NORMALIZE_MFG.items():
        if canonical in NORMALIZE_MFG and NORMALIZE_MFG[canonical] != canonical:
            circular.append(f"{abbrev} → {canonical} → {NORMALIZE_MFG[canonical]}")

    log.result("No circular references", len(circular) == 0, f"{len(circular)} found" if circular else "Clean")

    # Verify critical normalizations
    critical_cases = {
        'CROUSE-HINDS': 'CROUSE HINDS',
        'STATIC O RING': 'STATIC O-RING',
        'PANDT': 'PANDUIT',
    }

    for input_val, expected in critical_cases.items():
        result = NORMALIZE_MFG.get(input_val)
        log.result(f"Canonical: {input_val} → {expected}", result == expected, f"got={result}")

    log.result("Distributors list populated", len(DISTRIBUTORS) >= 4, f"{len(DISTRIBUTORS)} defined")
    log.result("Descriptors list populated", len(DESCRIPTORS) >= 5, f"{len(DESCRIPTORS)} defined")

    return {
        'normalize_map_size': len(NORMALIZE_MFG),
        'distributors_count': len(DISTRIBUTORS),
        'descriptors_count': len(DESCRIPTORS)
    }

# ═══════════════════════════════════════════════════════════════
# TEST PHASE 6: EDGE CASES — Boundary Conditions
# ═══════════════════════════════════════════════════════════════

def test_edge_cases(log):
    """Test parser behavior on edge cases and boundary conditions."""
    log.section("PHASE 6: EDGE CASES — Boundary Conditions")

    # Null and empty inputs
    pn, method = extract_pn_from_text(None)
    log.result("PN extract: None input", pn is None, f"got={pn}")

    pn, method = extract_pn_from_text("")
    log.result("PN extract: empty string", pn is None, f"got={pn}")

    mfg, method = extract_mfg_from_text(None, None, set())
    log.result("MFG extract: None input", mfg is None, f"got={mfg}")

    result = sanitize_mfg(None)
    log.result("Sanitize: None input", result is None, f"got={result}")

    result = sanitize_mfg("")
    log.result("Sanitize: empty string", result is None, f"got={result}")

    # Very long input
    long_text = "A" * 5000 + " MANUFACTURER: PANDUIT PN: ABC123 " + "B" * 5000
    pn, method = extract_pn_from_text(long_text)
    log.result("PN extract: 10K char input", pn is not None, f"got={pn}")

    # Special characters
    pn, method = extract_pn_from_text("PN: ABC-123/DEF_456.789")
    log.result("PN extract: special chars (-/_/.)", pn is not None, f"got={pn}")

    # Multiple labels (priority test)
    pn, method = extract_pn_from_text("PN: FIRST123 MODEL NUMBER: SECOND456")
    log.result("PN extract: multiple labels", pn == "FIRST123", f"got={pn} (should pick first)")

    return {}

# ═══════════════════════════════════════════════════════════════
# REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════

def generate_report(output_dir, all_results, log):
    """Generate Excel validation report with all test results."""
    log.section("GENERATING VALIDATION REPORT")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.warning("openpyxl not available - skipping Excel report generation")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    # Formatting
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='009639')  # Wesco green
    pass_fill = PatternFill('solid', fgColor='C6EFCE')
    fail_fill = PatternFill('solid', fgColor='FFC7CE')
    section_font = Font(bold=True, size=11)

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50

    # Header row
    row = 1
    for col, val in enumerate(['Metric', 'Result', 'Detail'], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    # Summary section
    row = 2
    ws.cell(row=row, column=1, value="Test Run").font = section_font
    ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

    accuracy = all_results.get('accuracy')
    if accuracy:
        row += 2
        ws.cell(row=row, column=1, value="ACCURACY TEST").font = section_font
        for field, label in [('mfg', 'MFG'), ('pn', 'PN')]:
            row += 1
            r = accuracy[field]
            ws.cell(row=row, column=1, value=f"{label} Exact Match")
            c = ws.cell(row=row, column=2, value=f"{r['accuracy_exact']}%")
            c.fill = pass_fill if r['accuracy_exact'] >= 85 else fail_fill
            ws.cell(row=row, column=3, value=f"{r['exact_match']}/{r['scorable_rows']} rows")

            row += 1
            ws.cell(row=row, column=1, value=f"{label} Fuzzy Match")
            c = ws.cell(row=row, column=2, value=f"{r['accuracy_fuzzy']}%")
            c.fill = pass_fill if r['accuracy_fuzzy'] >= 92 else fail_fill
            ws.cell(row=row, column=3, value="Includes partial matches")

    # Add test log sheet
    ws2 = wb.create_sheet("Test Log")
    ws2.append(['Test', 'Status', 'Detail', 'Timestamp'])
    for col in range(1, 5):
        ws2.cell(row=1, column=col).font = Font(bold=True)

    for entry in log.entries:
        status = 'PASS' if entry['passed'] else 'FAIL'
        ws2.append([entry['test'], status, entry['detail'], entry['timestamp']])
        row_num = ws2.max_row
        ws2.cell(row=row_num, column=2).fill = pass_fill if entry['passed'] else fail_fill

    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['C'].width = 60

    # Save report
    report_path = Path(output_dir) / "validation_report.xlsx"
    wb.save(str(report_path))
    log.info(f"Excel report saved: {report_path}")
    return str(report_path)

# ═══════════════════════════════════════════════════════════════
# MAIN EXECUTION
# ═══════════════════════════════════════════════════════════════

def main():
    """Main test runner."""
    parser = argparse.ArgumentParser(description="Wesco MRO Parser Validation Suite")
    parser.add_argument('--data-dir', default=DEFAULT_DATA_DIR, help="Directory containing test data files")
    parser.add_argument('--output-dir', default=None, help="Output directory for results (default: data-dir/test_results)")
    parser.add_argument('--quick', action='store_true', help="Run quick tests only (skip Book25)")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    output_dir = Path(args.output_dir) if args.output_dir else data_dir / "test_results"

    # Initialize logger
    log = TestLogger(output_dir)
    log.info(f"Data directory: {data_dir}")
    log.info(f"Output directory: {output_dir}")

    # Pre-flight checks
    log.section("PRE-FLIGHT CHECKS")
    for fname in [ELECTRICAL_BLANK, ELECTRICAL_COMPLETE, BOOK25]:
        exists = (data_dir / fname).exists()
        log.result(f"File exists: {fname}", exists)
        if not exists and fname != BOOK25:
            log.warning("Missing required test file - some tests will be skipped")

    # Run test phases
    all_results = {}

    all_results['accuracy'] = test_accuracy(data_dir, output_dir, log)
    all_results['qa'] = test_qa_engine(output_dir, log)
    all_results['unit'] = test_unit_extraction(log)

    if not args.quick:
        all_results['scale'] = test_scale_book25(data_dir, output_dir, log)

    all_results['normalization'] = test_normalization(log)
    all_results['edge_cases'] = test_edge_cases(log)

    # Generate reports
    generate_report(output_dir, all_results, log)

    # Save JSON results
    json_path = output_dir / "test_results.json"
    with open(json_path, 'w') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'results': {k: v for k, v in all_results.items() if v is not None},
            'test_entries': log.entries
        }, f, indent=2, default=str)
    log.info(f"JSON results: {json_path}")

    # Final summary
    passed, failed = log.summary()

    return 0 if failed == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
